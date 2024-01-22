import os
import openpyxl
import discord
from discord import app_commands
from discord.ext import commands
from openpyxl.styles import Alignment, Border, Side
from discord.ext import tasks
from datetime import datetime
from core.twitch import TwitchStreamChecker, get_name_from_url
from core.youtube import YTstats, get_channel_ids,API_KEY,get_channel_name,get_latest_ids, get_live_stream_id
import random
from datetime import datetime, timedelta, timezone


def convert_timestamp(timestamp):
    dt_object = datetime.strptime(timestamp, '%Y-%m-%dT%H:%M:%SZ')
    utc_dt = dt_object.replace(tzinfo=timezone.utc)
    
    riyadh_offset = timedelta(hours=3)
    riyadh_dt = utc_dt + riyadh_offset
    
    formatted_date = riyadh_dt.strftime('%m/%d/%Y %I:%M %p')
    
    return formatted_date



class stream_cog(commands.Cog):
    def __init__(self, bot: commands.Bot) -> None:
        self.bot = bot
        self.twitch_checker = TwitchStreamChecker()


    @app_commands.command(name="announcechannel", description="Sets the discord channel where stream announcements will be posted.")
    @app_commands.guild_only()
    async def announcechannel(self, interaction: discord.Interaction, channel: discord.abc.GuildChannel):
        channel_id = str(channel.id)  # Convert to string
        user_name = interaction.user.name
        guild = str(interaction.guild.id)  # Convert to string
        await interaction.response.defer(thinking=True)

        if not (interaction.user.guild_permissions.administrator or interaction.user.guild_permissions.manage_channels):
            embed = discord.Embed(
                title="Invalid user",
                description="This command requires the ``Administrator or Manage channels`` permission",
                color=discord.Color.red()
            )
            embed.set_author(name=f"OAN", icon_url="https://media.discordapp.net/attachments/1059135171540029552/1130471478526214194/1688931915684.png?width=451&height=660")
            if interaction.user.avatar:
                embed.set_footer(text=f'{user_name}', icon_url=interaction.user.avatar.url)
            await interaction.followup.send(embed=embed)
            return

        else:
            #Twitch
            sheet_path = os.path.join("core", "stream_guild.xlsx")
            wb = openpyxl.load_workbook(sheet_path)
            sheet_names = wb.sheetnames
            if guild not in sheet_names:
                # Create a new worksheet with the value of guild_id
                wb.create_sheet(guild)
                wb.create_sheet(guild+"_youtube")
            sheet = wb[guild]
                #SET THE DEFUALT CHANNEL
            if sheet.cell(1, 3).value is None:
                sheet.cell(1, 3).value = channel_id
                wb.save(sheet_path)
                sheet = wb[guild+"_youtube"]
                if sheet.cell(1, 3).value is None:
                    sheet.cell(1, 3).value = channel_id
                    wb.save(sheet_path)
                embed = discord.Embed(
                title=f"OAN default announcechannel is ``{channel.name}``",
                description=f"The {channel.name} has been set successfully\n Now you can add streamers by </addstreamer:1135989711253536789>",
                color=0x6a95a2
            )
                embed.set_author(name=f"OAN", icon_url="https://media.discordapp.net/attachments/1059135171540029552/1130471478526214194/1688931915684.png?width=451&height=660")
                if interaction.user.avatar:
                    embed.set_footer(text=f'{user_name}', icon_url=interaction.user.avatar.url)
                await interaction.followup.send(embed=embed)
                print(f"************************************************\n guild {interaction.guild.name} : {interaction.user.name}  has been set announcechannel\n **************************************")
                return


            elif channel_id == str(sheet.cell(1, 3).value):
                embed = discord.Embed(
                        title=f"Registration ``{channel}``",
                        description=f"**{channel}**  has already been registered",
                        color=0x6a95a2
                    )
                embed.set_author(name=f"OAN", icon_url="https://media.discordapp.net/attachments/1059135171540029552/1130471478526214194/1688931915684.png?width=451&height=660")
                if interaction.user.avatar:
                    embed.set_footer(text=f'{user_name}', icon_url=interaction.user.avatar.url)
                await interaction.followup.send(embed=embed)
                print(f"************************************************\n guild {interaction.guild.name} : {interaction.user.name} has already been registered announcechannel\n **************************************")
                return
                

            else:
                sheet.cell(1, 3).value = channel_id
                wb.save(sheet_path)
                sheet = wb[guild+"_youtube"]
                sheet.cell(1, 3).value = channel_id
                wb.save(sheet_path)
                embed = discord.Embed(
                    title=f"Channel updated: ``{channel.name}``",
                    description="This channel has been updated in the registration",
                    color=0x6a95a2
                )
                embed.set_author(name=f"OAN", icon_url="https://media.discordapp.net/attachments/1059135171540029552/1130471478526214194/1688931915684.png?width=451&height=660")
                if interaction.user.avatar:
                    embed.set_footer(text=f'{user_name}', icon_url=interaction.user.avatar.url)
                await interaction.followup.send(embed=embed)
                print(f"************************************************\n guild {interaction.guild.name} : {interaction.user.name} has been updated\n **************************************")
                return

    @app_commands.command(name="addstreamer", description="Adds a new streamer to list of streamers being monitored")
    @app_commands.choices(streaming_service=[
    app_commands.Choice(name='Twitch', value=1),
    app_commands.Choice(name='Youtube', value=2)])
    @app_commands.guild_only()
    async def addstreamers(self, interaction: discord.Interaction, streaming_service: app_commands.Choice[int],streamer: str, mention: str = None):
        try:
            user_name = interaction.user.name
            guild = str(interaction.guild.id)  # Convert to string
            await interaction.response.defer(thinking=True)
            if not (interaction.user.guild_permissions.administrator or interaction.user.guild_permissions.manage_channels):
                embed = discord.Embed(
                    title="Invalid user",
                    description="This command requires the ``Administrator or Manage channels`` permission",
                    color=discord.Color.red()
                )
                embed.set_author(name=f"OAN", icon_url="https://media.discordapp.net/attachments/1059135171540029552/1130471478526214194/1688931915684.png?width=451&height=660")
                if interaction.user.avatar:
                    embed.set_footer(text=f'{user_name}', icon_url=interaction.user.avatar.url)
                await interaction.followup.send(embed=embed)
                print(f"************************************************\n guild {interaction.guild.name} : {interaction.user.name} has already send you don't have permistion\n **************************************")

                return
            
            else:

                sheet_path = os.path.join("core", "stream_guild.xlsx")
                wb = openpyxl.load_workbook(sheet_path)
                sheet_names = wb.sheetnames
                if guild not in sheet_names:
                    embed = discord.Embed(
                        title="Invaild announcechannel",
                        description="Please use command </announcechannel:1135989711253536788> first to Set the discord channel where stream announcements will be posted.",
                        color=discord.Color.dark_orange()
                    )
                    embed.set_author(name=f"OAN", icon_url="https://media.discordapp.net/attachments/1059135171540029552/1130471478526214194/1688931915684.png?width=451&height=660")
                    if interaction.user.avatar:
                        embed.set_footer(text=f'{user_name}', icon_url=interaction.user.avatar.url)
                    await interaction.followup.send(embed=embed)
                    print(f"************************************************\n guild {interaction.guild.name} : {interaction.user.name} has already send use announcements first\n **************************************")
                    return


                #check which service 
                #Twitch
                if streaming_service.value == 1:
                    sheet = wb[guild]
                    max_rows = sheet.max_row + 1
                    streamer = get_name_from_url(streamer)
                    for row in range(1, max_rows):  # Start from row 2
                        if str(streamer).lower() == str(sheet.cell(row, 1).value).lower():
                            embed = discord.Embed(
                                title=f"``{streamer}'s registration``",
                                description=f"**{streamer}**  has already been registered by {str(sheet.cell(row=row, column=4).value)}",
                                color=0x6a95a2
                            )
                            embed.set_author(name=f"OAN", icon_url="https://media.discordapp.net/attachments/1085541383563124858/1135948796006781008/-51609794436soaav8dnrc.png?width=537&height=458")
                            if interaction.user.avatar:
                                embed.set_footer(text=f'{user_name}', icon_url=interaction.user.avatar.url)
                            await interaction.followup.send(embed=embed)
                            print(f"************************************************\n guild {interaction.guild.name} has already been registered addstreamers\n **************************************")
                            sheet.cell(row, 5).value = mention
                            wb.save(sheet_path)
                            return
                        
                    check_streamer = TwitchStreamChecker()
                    if check_streamer.is_user_exist(streamer.lower()) :
                        alignment = Alignment(horizontal='center', vertical='center')
                        cell = sheet.cell(max_rows, 1)
                        cell.alignment = alignment
                        # Apply bordsers and set color to gray
                        border = Border(
                            left=Side(border_style='thin', color='FF808080'),
                            right=Side(border_style='thin', color='FF808080'),
                            top=Side(border_style='thin', color='FF808080'),
                            bottom=Side(border_style='thin', color='FF808080')
                        )
                        cell.border = border
                        #sreamer name
                        cell = sheet.cell(max_rows, 1)
                        cell.alignment = alignment
                        cell.border = border
                        cell.value = streamer
                        #status
                        cell = sheet.cell(max_rows, 2)
                        cell.alignment = alignment
                        cell.border = border
                        cell.value = "False"
                        #---------------------
                        #guild_name
                        cell = sheet.cell(max_rows, 3)
                        cell.alignment = alignment
                        cell.border = border
                        cell.value = interaction.guild.name
                        #author
                        cell = sheet.cell(max_rows, 4)
                        cell.alignment = alignment
                        cell.border = border
                        cell.value = interaction.user.name
                        #message
                        cell = sheet.cell(max_rows, 5)
                        cell.alignment = alignment
                        cell.border = border
                        cell.value = mention
                        wb.save(sheet_path)
                        embed = discord.Embed(
                            title=f"``{streamer}'s`` registration",
                            description=f"**{streamer}** has been set successfully",
                            color=0x6a95a2
                        )
                        embed.set_author(name=f"OAN", icon_url="https://media.discordapp.net/attachments/1085541383563124858/1135948796006781008/-51609794436soaav8dnrc.png?width=537&height=458")
                        if interaction.user.avatar:
                            embed.set_footer(text=f'{user_name}', icon_url=interaction.user.avatar.url)
                        await interaction.followup.send(embed=embed)
                        print(f"************************************************\n guild {interaction.guild.name} : {interaction.user.name} has been set add streamer \n **************************************")
                        return
                    else: 
                        embed = discord.Embed(
                        title=f"``{streamer}'s``  registration",
                        description=f"**{streamer}** Can't find the user in Twitch please check the name!",
                        color=discord.Color.red()
                        )
                        embed.set_author(name=f"OAN", icon_url="https://media.discordapp.net/attachments/1085541383563124858/1135948796006781008/-51609794436soaav8dnrc.png?width=537&height=458")

                        if interaction.user.avatar:
                            embed.set_footer(text=f'{user_name}', icon_url=interaction.user.avatar.url)
                        await interaction.followup.send(embed=embed)
                        print(f"************************************************\n guild {interaction.guild.name} : {interaction.user.name} Can't find the user in Twitch \n **************************************")
                        return
                #Youtube
                elif streaming_service.value == 2:
                    streamer_id = get_channel_ids(streamer)
                    sheet = wb[guild+"_youtube"]
                    max_rows = sheet.max_row + 1
                    if streamer_id:
                        for row in range(1, max_rows):  # Start from row 2
                            if str(streamer_id).lower() == str(sheet.cell(row, 2).value).lower():
                                embed = discord.Embed(
                                    title=f"``{streamer}'s registration``",
                                    description=f"**{streamer}**  has already been registered by {str(sheet.cell(row=row, column=5).value)}",
                                    color=0x6a95a2
                                )
                                embed.set_author(name=f"OAN", icon_url="https://images-ext-1.discordapp.net/external/Rzj_Aw6pkT9xxO1_yjYycxnXn0S0LQMSj47KZEpVVJM/https/cdn.pixabay.com/photo/2021/09/11/18/21/youtube-6616310_960_720.png?width=921&height=660")
                                if interaction.user.avatar:
                                    embed.set_footer(text=f'{user_name}', icon_url=interaction.user.avatar.url)
                                await interaction.followup.send(embed=embed)
                                print(f"************************************************\n guild {interaction.guild.name} has already been registered addstreamers\n **************************************")
                                sheet.cell(row, 6).value = mention
                                wb.save(sheet_path)
                                return
                            
                    if streamer_id:
                        alignment = Alignment(horizontal='center', vertical='center')
                        cell = sheet.cell(max_rows, 1)
                        cell.alignment = alignment
                        # Apply bordsers and set color to gray
                        border = Border(
                            left=Side(border_style='thin', color='FF808080'),
                            right=Side(border_style='thin', color='FF808080'),
                            top=Side(border_style='thin', color='FF808080'),
                            bottom=Side(border_style='thin', color='FF808080')
                        )
                        cell.border = border
                        #sreamer ID
                        cell = sheet.cell(max_rows, 2)
                        cell.alignment = alignment
                        cell.border = border
                        cell.value = streamer_id
                        #sreamer name
                        cell = sheet.cell(max_rows, 1)
                        cell.alignment = alignment
                        cell.border = border
                        name_channel = get_channel_name(streamer)
                        cell.value = name_channel
                        #---------------------
                        #guild_name
                        cell = sheet.cell(max_rows, 4)
                        cell.alignment = alignment
                        cell.border = border
                        cell.value = interaction.guild.name
                        #author
                        cell = sheet.cell(max_rows, 5)
                        cell.alignment = alignment
                        cell.border = border
                        cell.value = interaction.user.name
                        #message
                        cell = sheet.cell(max_rows, 6)
                        cell.alignment = alignment
                        cell.border = border
                        cell.value = mention
                        #id_vides
                        youtube = YTstats(random.choice(API_KEY), streamer_id)
                        channel_videos= youtube.get_channel_video_data()
                        statistics = youtube.get_channel_statistics()
                        channel_videos_str = ",".join(channel_videos)
                        cell = sheet.cell(max_rows, 7)
                        cell.alignment = alignment
                        cell.border = border
                        cell.value = channel_videos_str
                        wb.save(sheet_path)
                        embed = discord.Embed(
                            title=f"``{streamer}'s`` registration",
                            description=f"**{streamer}** has been set successfully",
                            color=0x6a95a2
                        )
                        embed.set_thumbnail(url=statistics['profile_image_url_high'])
                        embed.set_author(name=f"OAN", icon_url="https://images-ext-1.discordapp.net/external/Rzj_Aw6pkT9xxO1_yjYycxnXn0S0LQMSj47KZEpVVJM/https/cdn.pixabay.com/photo/2021/09/11/18/21/youtube-6616310_960_720.png?width=921&height=660")
                        embed.add_field(name="Channel ID", value=f"`{streamer_id}`",inline=True)
                        embed.add_field(name="View Count", value=f"`{statistics['statistics']['viewCount']}`",inline=False)
                        embed.add_field(name="Subscriber Count", value=f"`{statistics['statistics']['subscriberCount']}`",inline=True)
                        embed.add_field(name="Video Count", value=f"`{statistics['statistics']['videoCount']}`",inline=False)
                        if interaction.user.avatar:
                            embed.set_footer(text=f'{user_name}', icon_url=interaction.user.avatar.url)
                        await interaction.followup.send(embed=embed)
                        print(f"************************************************\n guild {interaction.guild.name} : {interaction.user.name} has been set add streamer \n **************************************")
                        return
                    else: 
                        embed = discord.Embed(
                        title=f"``{streamer}'s``  registration",
                        description=f"**{streamer}** Can't find the user in Youtube please check the name!",
                        color=discord.Color.red()
                        )
                        embed.set_author(name=f"OAN", icon_url="https://images-ext-1.discordapp.net/external/Rzj_Aw6pkT9xxO1_yjYycxnXn0S0LQMSj47KZEpVVJM/https/cdn.pixabay.com/photo/2021/09/11/18/21/youtube-6616310_960_720.png?width=921&height=660")

                        if interaction.user.avatar:
                            embed.set_footer(text=f'{user_name}', icon_url=interaction.user.avatar.url)
                        await interaction.followup.send(embed=embed)
                        print(f"************************************************\n guild {interaction.guild.name} : {interaction.user.name} Can't find the user in Twitch \n **************************************")
                        return


                
        except Exception as e : 
            print(e)

    @app_commands.command(name="streamers", description="Shows a list of streamers currently being monitored.")
    @app_commands.choices(streaming_service=[
    app_commands.Choice(name='Twitch', value=1),
    app_commands.Choice(name='Youtube', value=2)])
    @app_commands.guild_only()
    async def streamers(self, interaction: discord.Interaction, streaming_service: app_commands.Choice[int]):
        try:
            user_name = interaction.user.name
            guild = str(interaction.guild.id)  # Convert to string
            await interaction.response.defer(thinking=True)

            if not (interaction.user.guild_permissions.administrator or interaction.user.guild_permissions.manage_channels):
                embed = discord.Embed(
                    title="Invalid user",
                    description="This command requires the ``Administrator or Manage channels`` permission",
                    color=discord.Color.red()
                )
                embed.set_author(name=f"OAN", icon_url="https://media.discordapp.net/attachments/1059135171540029552/1130471478526214194/1688931915684.png?width=451&height=660")
                if interaction.user.avatar:
                    embed.set_footer(text=f'{user_name}', icon_url=interaction.user.avatar.url)
                print(f"************************************************\n guild {interaction.guild.name} : {interaction.user.name} has already send you don't have permistion\n **************************************")

                await interaction.followup.send(embed=embed)

                return

            else:
                sheet_path = os.path.join("core", "stream_guild.xlsx")
                wb = openpyxl.load_workbook(sheet_path)
                sheet_names = wb.sheetnames
                if guild not in sheet_names:
                    embed = discord.Embed(
                        title="Invaild announcechannel",
                        description="Please use command </announcechannel:1135989711253536788> first to Sets the discord channel where stream announcements will be posted.",
                        color=discord.Color.dark_orange()
                    )
                    embed.set_author(name=f"OAN", icon_url="https://media.discordapp.net/attachments/1059135171540029552/1130471478526214194/1688931915684.png?width=451&height=660")
                    if interaction.user.avatar:
                        embed.set_footer(text=f'{user_name}', icon_url=interaction.user.avatar.url)
                    await interaction.followup.send(embed=embed)
                    print(f"************************************************\n guild {interaction.guild.name} : {interaction.user.name} has already send use announcechannel\n **************************************")
                    return
                if streaming_service.value ==1: 
                    sheet = wb[guild]
                else:                     
                    sheet = wb[guild+"_youtube"]
                streamers = []
                max_rows = sheet.max_row + 1
                embed = discord.Embed(
                title=f"Streamers",
                description=f"**list of streamers currently being monitored**",
                color=0x6a95a2)
                embed.set_author(name=f"OAN", icon_url="https://media.discordapp.net/attachments/1059135171540029552/1130471478526214194/1688931915684.png?width=451&height=660") 

                for row in range(2, max_rows): # St art from row 2
                    streamers.append(str(sheet.cell(row, 1).value))
                for streamer in streamers:
                    embed.add_field(name="Name", value=f"`{streamer}`", inline="False")
                if interaction.user.avatar:
                    embed.set_footer(text=f'{user_name}', icon_url=interaction.user.avatar.url)
                await interaction.followup.send(embed=embed)
                print(f"************************************************\n guild {interaction.guild.name} : {interaction.user.name} has send list of streamers\n **************************************")
                return

        except Exception as e : 
            print(e)

    @app_commands.command(name="removestreamer", description="Removes streamer from the  list of streamers being monitored")
    @app_commands.guild_only()
    @app_commands.choices(streaming_service=[
    app_commands.Choice(name='Twitch', value=1),
    app_commands.Choice(name='Youtube', value=2)])
    async def removestreamer(self, interaction: discord.Interaction,streaming_service: app_commands.Choice[int], removestreamer: str):
        try:
            user_name = interaction.user.name
            guild = str(interaction.guild.id)  # Convert to string
            await interaction.response.defer(thinking=True)

            if not (interaction.user.guild_permissions.administrator or interaction.user.guild_permissions.manage_channels):
                embed = discord.Embed(
                    title="Invalid user",
                    description="This command requires the ``Administrator or Manage channels`` permission",
                    color=discord.Color.red()
                )
                embed.set_author(name=f"OAN", icon_url="https://media.discordapp.net/attachments/1059135171540029552/1130471478526214194/1688931915684.png?width=451&height=660")
                if interaction.user.avatar:
                    embed.set_footer(text=f'{user_name}', icon_url=interaction.user.avatar.url)
                print(f"************************************************\n guild {interaction.guild.name} : {interaction.user.name} has already send you don't have permistion\n **************************************")

                await interaction.followup.send(embed=embed)

                return

            else:
                sheet_path = os.path.join("core", "stream_guild.xlsx")
                wb = openpyxl.load_workbook(sheet_path)
                sheet_names = wb.sheetnames
                if guild not in sheet_names:
                    embed = discord.Embed(
                        title="Invaild Streamer",
                        description="You have to set </announcechannel:1135989711253536788> first! ",
                        color=discord.Color.dark_orange()
                    )
                    embed.set_author(name=f"OAN", icon_url="https://media.discordapp.net/attachments/1059135171540029552/1130471478526214194/1688931915684.png?width=451&height=660")

                    if interaction.user.avatar:
                        embed.set_footer(text=f'{user_name}', icon_url=interaction.user.avatar.url)
                    await interaction.followup.send(embed=embed)
                    print(f"************************************************\n guild {interaction.guild.name} : {interaction.user.name} has already send use announcechannel\n **************************************")
                    return
                
                if streaming_service.value ==1: 
                    sheet = wb[guild]
                    streamers = []
                    max_rows = sheet.max_row + 1 
                    remove_row_index = None
                    for row in range(2, max_rows): # Start from row 2
                        streamer = str(sheet.cell(row, 1).value).lower()
                        streamers.append(streamer)
                        if streamer == removestreamer.lower():
                            remove_row_index = row
                else:                     
                    sheet = wb[guild+"_youtube"]
                    streamers = []
                    max_rows = sheet.max_row + 1 
                    remove_row_index = None
                    id_removestreamer = get_channel_ids(removestreamer)
                    for row in range(2, max_rows): # Start from row 2
                        streamer = str(sheet.cell(row, 2).value).lower()
                        streamers.append(streamer)
                        try:
                            if streamer == id_removestreamer.lower():
                                remove_row_index = row
                        except Exception:
                            pass



                # If removestreamer is found, delete the row
                if remove_row_index is not None:
                    sheet.delete_rows(remove_row_index)
                    embed = discord.Embed(
                    title=f"Remove Streamer",
                    description=f"**`{removestreamer}` has been deleted successfully**",
                    color=0x6a95a2)
                    embed.set_author(name=f"OAN", icon_url="https://media.discordapp.net/attachments/1059135171540029552/1130471478526214194/1688931915684.png?width=451&height=660")
                    if interaction.user.avatar:
                        embed.set_footer(text=f'{user_name}', icon_url=interaction.user.avatar.url)
                    await interaction.followup.send(embed=embed)
                    wb.save(sheet_path)
                    print(f"************************************************\n guild {interaction.guild.name}: {interaction.user.name} has been deleted {removestreamer}\n **************************************")
                    return
                else:
                    embed = discord.Embed(
                    title=f"Remove Streamer",
                    description=f"**`{removestreamer}` Invaild** you can know the streamers using </streamers:1135989711253536790> ",
                    color=discord.Color.dark_orange())
                    embed.set_author(name=f"OAN", icon_url="https://media.discordapp.net/attachments/1059135171540029552/1130471478526214194/1688931915684.png?width=451&height=660")
                    if interaction.user.avatar:
                        embed.set_footer(text=f'{user_name}', icon_url=interaction.user.avatar.url)
                    await interaction.followup.send(embed=embed)
                    print(f"************************************************\n guild {interaction.guild.name}: {interaction.user.name} has been try to delete invalid {removestreamer}\n **************************************")
                    return

        except Exception as e : 
            print(e)


                            



async def setup(bot: commands.Bot) -> None:
    await bot.add_cog(stream_cog(bot))


