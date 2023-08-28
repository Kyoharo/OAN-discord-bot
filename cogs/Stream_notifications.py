
import os
import openpyxl
import discord
from discord import app_commands
from discord.ext import commands
from openpyxl.styles import Alignment, Border, Side
from discord.ext import tasks
from datetime import datetime
from core.twitch import TwitchStreamChecker, get_name_from_url
from core.youtube import YTstats, get_channel_ids,API_KEY,get_channel_name,get_latest_ids, get_live_stream_id
import random
from datetime import datetime, timedelta, timezone


def convert_timestamp(timestamp):
    # Convert the input timestamp to a datetime object
    dt_object = datetime.strptime(timestamp, '%Y-%m-%dT%H:%M:%SZ')
    
    # Convert the datetime object to UTC time (assuming the input timestamp is in UTC)
    utc_dt = dt_object.replace(tzinfo=timezone.utc)
    
    # Convert to Riyadh time (UTC+3)
    riyadh_offset = timedelta(hours=3)
    riyadh_dt = utc_dt + riyadh_offset
    
    # Convert the Riyadh datetime to the desired format
    formatted_date = riyadh_dt.strftime('%m/%d/%Y %I:%M %p')
    
    return formatted_date



class stream_notifications(commands.Cog):
    def __init__(self, bot: commands.Bot) -> None:
        self.bot = bot
        self.twitch_checker = TwitchStreamChecker()
        self.update_stream_status.start()
        self.update_youtube_status.start()
        self.update_youtube_live.start()

    @tasks.loop(seconds=20)
    async def update_stream_status(self):
        sheet_path = os.path.join("core", "stream_guild.xlsx")
        wb = openpyxl.load_workbook(sheet_path)
        sheet_names = wb.sheetnames
        for guild in sheet_names:
            if guild.endswith("youtube"):
                continue
            sheet = wb[guild]
            max_rows = sheet.max_row + 1
            for row in range(2, max_rows): 
                streamer = str(sheet.cell(row, 1).value)
                stream_info = self.twitch_checker.is_streamer_live(streamer)
                if stream_info['is_live']:
                    status = str(sheet.cell(row, 2).value)
                    if status != "True":
                        channel_id = int(sheet.cell(1, 3).value)
                        message = str(sheet.cell(row, 5).value)
                        text_channel = self.bot.get_channel(channel_id)
                        embed = discord.Embed(
                            title=stream_info['stream_title'],  # Use the stream title directly as the title
                            color=0x6a95a2,
                            url=f"https://www.twitch.tv/{streamer}"  # Set the URL for the title hyperlink
                        )
                        embed.set_author(name=f"[LIVE] {streamer}", icon_url=stream_info['profile_image_url'], url=f"https://www.twitch.tv/{streamer}") 
                        embed.set_thumbnail(url=stream_info['profile_image_url'])
                
                        width = 500
                        height = 250
                        # Modify the thumbnail_url to use the desired width and height
                        thumbnail_url = stream_info['thumbnail_url'].replace("{width}", str(width)).replace("{height}", str(height))
                        embed.set_image(url=thumbnail_url)
                        embed.add_field(name="Game", value=stream_info['game_name'])
                        embed.add_field(name="Viewer Count", value=stream_info['viewer_count'])
                        # started_at_utc = datetime.strptime(stream_info['started_at'], "%Y-%m-%dT%H:%M:%SZ")
                        #add button
                        view = discord.ui.View()
                        button = discord.ui.Button(emoji="<:Twitch:1136322224018694195>]",label=f" Twitch.com/{streamer}",url=f"https://www.twitch.tv/{streamer}")
                        view.add_item(button)
                        # Convert to the desired format
                        time_format = convert_timestamp(stream_info['started_at'])

                        embed.set_footer(icon_url="https://media.discordapp.net/attachments/1085541383563124858/1135948796006781008/-51609794436soaav8dnrc.png?width=537&height=458",text=f"Twitch {time_format}")

                        if str(message) == 'None':
                            regular_message = f"**{streamer} is now LIVE on Twitch @everyone**"
                        else:
                            regular_message = f"**{streamer} is now LIVE on Twitch {message}**"

                        await text_channel.send(regular_message, embed=embed,view=view)
                        print(f"\n\n twitch {streamer} have been sended \n\n")                    
                    sheet.cell(row, 2).value = "True"
                    wb.save(sheet_path)
                else:
                    sheet.cell(row, 2).value = "False"
                    wb.save(sheet_path)


    @tasks.loop(seconds=20)
    async def update_youtube_status(self):                      ##Youtube videos

        sheet_path = os.path.join("core", "stream_guild.xlsx")
        wb = openpyxl.load_workbook(sheet_path)
        sheet_names = wb.sheetnames
        for guild in sheet_names:
            if not guild.endswith("youtube"):
                continue
            sheet = wb[guild]
            max_rows = sheet.max_row + 1
            for row in range(2, max_rows): 
                streamer = str(sheet.cell(row, 1).value)
                last_ids = get_latest_ids(streamer)
                previous_id =  str(sheet.cell(row, 7).value)
                try:
                    previous_id = previous_id.split(',')
                except Exception: 
                    pass
                # print(f"previous_id : {previous_id}  last_ids: {last_ids}")
                if all(id in previous_id for id in last_ids):
                    # print("All items are present")
                    pass
                else:
                    # print("Not all items are present")
                    streamer_id = str(sheet.cell(row, 2).value)
                    youtube = YTstats(random.choice(API_KEY), streamer_id)
                    try:
                        channel_videos= youtube.get_channel_video_data()
                    except Exception as e : 
                        youtube = YTstats(random.choice(API_KEY), streamer_id)
                        channel_videos= youtube.get_channel_video_data()

                    # print(f"last new 5 ids : {channel_videos}   last scrap 2 ids : {last_ids}  ")
                    # if one item from the scrapping in the last 5 videos 
                    new_ids = [id for id in channel_videos if id not in previous_id]
                    if new_ids:
                        statistics = youtube.get_channel_statistics()
                                            
                        for video_id in new_ids:
                            # Fetch additional data for each video using YouTube API
                            video_data = youtube._get_video_data(video_id)
                            # print("Video data:", video_data)
                            embed = discord.Embed(
                                                        title=video_data['title'],  # Use the stream title directly as the title
                                                        color=0x6a95a2,
                                                        url=f"https://www.youtube.com/watch?v={video_id}"  # Set the URL for the title hyperlink
                                                    )
                            embed.set_author(name=f"{statistics['username']}", icon_url=statistics['profile_image_url_high'], url=f"https://www.youtube.com/@{streamer}") 
                                            
                            width = 500
                            height = 250
                            embed.set_thumbnail(url=statistics['profile_image_url_high'])
                            # Modify the thumbnail_url to use the desired width and height
                            video_data['thumbnails_high'] = video_data['thumbnails_high'].replace("/hqdefault.jpg", "/maxresdefault.jpg")
                            thumbnail_url = video_data['thumbnails_high'].replace("{width}", str(width)).replace("{height}", str(height))
                            embed.set_image(url=thumbnail_url)
                            description = video_data['description'].split('\n')[0]
                            embed.add_field(name="Description", value=description)
                            time_format = convert_timestamp(video_data['publishedAt'])

                            view = discord.ui.View()
                            button = discord.ui.Button(emoji="<:youtube:1142052855918895104>]",label=f" Youtube.com/{statistics['username']}",url=f"https://www.youtube.com/watch?v={video_id}")
                            view.add_item(button)
                            embed.set_footer(icon_url="https://images-ext-1.discordapp.net/external/Rzj_Aw6pkT9xxO1_yjYycxnXn0S0LQMSj47KZEpVVJM/https/cdn.pixabay.com/photo/2021/09/11/18/21/youtube-6616310_960_720.png?width=921&height=660",text=f"YouTube• {time_format}")
                            channel_id = int(sheet.cell(1, 3).value)
                            message = str(sheet.cell(row, 6).value)
                            text_channel = self.bot.get_channel(channel_id)
                            if str(message) == 'None':
                                regular_message = f"**@everyone {statistics['username']} just uploaded {video_data['title']}**"
                            else:
                                regular_message = f"**{message} {statistics['username']} just uploaded {video_data['title']}**"

                            await text_channel.send(regular_message, embed=embed,view=view)
                            print(f"\n\n Youtube video {streamer} have been sended \n\n")                    
                            # print("Appending IDs:", video_id)
                            previous_id.extend([video_id]) 
                            sheet.cell(row, 7, ','.join(previous_id))  # Update the cell value in the sheet
                    else:
                        # print("Append old video")
                        # Find the IDs that are in last_ids but not in previous_id
                        ids_to_append = [id for id in last_ids if id not in previous_id]
                        if ids_to_append:
                            # print("Appending IDs:", ids_to_append)
                            previous_id.extend(ids_to_append)  # Add the new IDs to previous_id list
                            sheet.cell(row, 7, ','.join(previous_id))  # Update the cell value in the sheet

        wb.save(sheet_path)  # Save the changes to the Excel shee

#-----------------------------------------------------------------------------------------------------------------------------------------------

    @tasks.loop(seconds=20)
    async def update_youtube_live(self):                      ##Youtube Live
        sheet_path = os.path.join("core", "stream_guild.xlsx")
        wb = openpyxl.load_workbook(sheet_path)
        sheet_names = wb.sheetnames
        for guild in sheet_names:
            if not guild.endswith("youtube"):
                continue
            sheet = wb[guild]
            max_rows = sheet.max_row + 1
            for row in range(2, max_rows): 
                streamer = str(sheet.cell(row, 1).value)
                previous_id =  str(sheet.cell(row, 3).value)
                live_id = get_live_stream_id(streamer)
                #all items form scraping are in the sheet
                try:
                    previous_id = previous_id.split(',')
                except Exception: 
                    pass
                #all items form scraping are in the sheet
                if live_id in previous_id or live_id == None:
                    pass
                    # print("Nothing new to send")
                else:
                    # print("gonna check the current status ")
                    streamer_id = str(sheet.cell(row, 2).value)
                    youtube = YTstats(random.choice(API_KEY), streamer_id)
                    streamer_live_id, stream_stautus = youtube.is_channel_live()
                    # print(f"streamer_live_id: {streamer_live_id} stream_stautus: {stream_stautus} ")
                    if stream_stautus == None:
                        # print("none")
                        previous_id.append(live_id)  # Append the new ID to the list
                        sheet.cell(row, 3, ','.join(previous_id))  # Update the cell value in the sheet
                    else:
                        # print(f"{streamer_live_id}  {stream_stautus}")
                        statistics = youtube.get_channel_statistics()
                        embed = discord.Embed(
                                                    title=stream_stautus['title'],  # Use the stream title directly as the title
                                                    color=0x6a95a2,
                                                    url=f"https://www.youtube.com/watch?v={streamer_live_id}"  # Set the URL for the title hyperlink
                                                )
                        embed.set_author(name=f"[LIVE] {statistics['username']}", icon_url=statistics['profile_image_url_high'], url=f"https://www.youtube.com/watch?v={streamer_live_id}") 
                                        
                        width = 500
                        height = 250
                        embed.set_thumbnail(url=statistics['profile_image_url_high'])
                        # Modify the thumbnail_url to use the desired width and height
                        stream_stautus['thumbnails_high'] = stream_stautus['thumbnails_high'].replace("/hqdefault.jpg", "/maxresdefault.jpg")
                        thumbnail_url = stream_stautus['thumbnails_high'].replace("{width}", str(width)).replace("{height}", str(height))
                        embed.set_image(url=thumbnail_url)
                        embed.add_field(name="Description", value=stream_stautus['description'])
                        time_format = convert_timestamp(stream_stautus['publishedAt'])

                        view = discord.ui.View()
                        button = discord.ui.Button(emoji="<:youtube:1142052855918895104>]",label=f" Youtube.com/{statistics['username']}",url=f"https://www.youtube.com/watch?v={streamer_live_id}")
                        view.add_item(button)
                        embed.set_footer(icon_url="https://images-ext-1.discordapp.net/external/Rzj_Aw6pkT9xxO1_yjYycxnXn0S0LQMSj47KZEpVVJM/https/cdn.pixabay.com/photo/2021/09/11/18/21/youtube-6616310_960_720.png?width=921&height=660",text=f"YouTube• {time_format}")
                        channel_id = int(sheet.cell(1, 3).value)
                        message = str(sheet.cell(row, 6).value)
                        text_channel = self.bot.get_channel(channel_id)
                        if str(message) == 'None':
                            regular_message = f"**@everyone {statistics['username']} is live Now!**"
                        else:
                            regular_message = f"**{message} {statistics['username']} is live Now!**"
                        await text_channel.send(regular_message, embed=embed,view=view)
                        print(f"\n\n Youtube LIVE {streamer} have been sended \n\n")                    
                        previous_id.append(streamer_live_id)  # Append the new ID to the list
                        sheet.cell(row, 3, ','.join(previous_id))  # Update the cell value in the sheet
                        previous_id_col7 = str(sheet.cell(row, 7).value).split(',')
                        previous_id_col7.append(streamer_live_id)
                        sheet.cell(row, 7, ','.join(previous_id_col7))  # Update the cell value in column 7

            wb.save(sheet_path)  # Save the changes to the Excel shee


    @update_youtube_live.before_loop 
    @update_youtube_status.before_loop
    @update_stream_status.before_loop   
    async def before_update_stream_status(self):
        await self.bot.wait_until_ready()


                            



async def setup(bot: commands.Bot) -> None:
    await bot.add_cog(stream_notifications(bot))


