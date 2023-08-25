import discord
from discord import app_commands
from discord.ext import commands
import openpyxl
from openpyxl.styles import Alignment, Border, Side
import os

class MyCog(commands.Cog):
    def __init__(self, bot: commands.Bot) -> None:
        self.bot = bot

    @app_commands.command(name="channel", description="Set a custom channel for the bot to work in.")
    @app_commands.guild_only()
    async def setchannel(self, interaction: discord.Interaction, channel: discord.TextChannel):
        channel_id = str(channel.id)  # Convert to string
        user_name = interaction.user.name
        guild = str(interaction.guild.id)  # Convert to string
        await interaction.response.defer(thinking=True)
        print("hello")
        if not (interaction.user.guild_permissions.administrator or interaction.user.guild_permissions.manage_channels):
            embed = discord.Embed(
                title="Invalid user",
                description="This command requires the ``Administrator`` permission",
                color=discord.Color.red()
            )
            if interaction.user.avatar:
                embed.set_footer(text=f'{user_name}', icon_url=interaction.user.avatar.url)
            await interaction.followup.send(embed=embed)
            return
        else:
            sheet_path = os.path.join("core", "guild.xlsx")
            wb = openpyxl.load_workbook(sheet_path)
            sheet = wb['Sheet1']
            max_rows = sheet.max_row + 1
            gg = sheet.cell(2, 1).value
            is_registered = False  # Flag variable
            
        for row in range(2, max_rows):  # Start from row 2
            if guild == str(sheet.cell(row, 1).value):
                if channel_id == str(sheet.cell(row, 2).value):
                    is_registered = True
                    break  # Channel is already registered, break the loop
                else:
                    sheet.cell(row, 2).value = channel_id
                    sheet.cell(row, 4).value = channel.name
                    is_registered = True
                    break
            
            if is_registered:
                embed = discord.Embed(
                    title=f"Channel updated: ``{channel.name}``",
                    description="This channel has been updated in the registration",
                    color=discord.Color.purple()
                )
                if interaction.user.avatar:
                    embed.set_footer(text=f'{user_name}', icon_url=interaction.user.avatar.url)
                await interaction.followup.send(embed=embed)
                print(f"************************************************\n guild {interaction.guild.name} has been updated\n **************************************")
                wb.save(sheet_path)
                return
            
            alignment = Alignment(horizontal='center', vertical='center')
            cell = sheet.cell(max_rows, 1)
            cell.alignment = alignment
            # Apply borders and set color to gray
            border = Border(
                left=Side(border_style='thin', color='FF808080'),
                right=Side(border_style='thin', color='FF808080'),
                top=Side(border_style='thin', color='FF808080'),
                bottom=Side(border_style='thin', color='FF808080')
            )
            cell.border = border
            cell.value = guild
            #---------------------
            #channel_id
            cell = sheet.cell(max_rows, 2)
            cell.alignment = alignment
            cell.border = border
            cell.value = channel_id
            #---------------------
            #guild name
            cell = sheet.cell(max_rows, 3)
            cell.alignment = alignment
            cell.border = border
            cell.value = interaction.guild.name
            #---------------------
            #channel name
            cell = sheet.cell(max_rows, 4)
            cell.alignment = alignment
            cell.border = border
            cell.value = channel.name
            wb.save(sheet_path)
            embed = discord.Embed(
                title=f"OAN default channel is now ``{channel.name}``",
                description="The channel has been set successfully",
                color=discord.Color.green()
            )
            if interaction.user.avatar:
                embed.set_footer(text=f'{user_name}', icon_url=interaction.user.avatar.url)
            await interaction.followup.send(embed=embed)
            try:
                print(f"************************************************\n guild {interaction.guild.name} has been set\n **************************************")
            except Exception as e:
                print(f"************************************************\n  {interaction.user.name} has been try to use setchannel\n **************************************")
            return

    @app_commands.command(name="info", description="Information about OAN Chatbot")
    async def info(self, interaction: discord.Interaction):
        user_name = interaction.user.name
        await interaction.response.defer(thinking=True)
        embed = discord.Embed(
            title="INFO",
            description="***OAN Artificial Intelligence Bot coded by Kyoharo <:Dev:1130428491620421652>***\nIntroducing **OAN**, the ultimate Discord bot that does it all! Chat in 100+ languages, share images and links effortlessly, track records, and even draw incredible images from your descriptions. Perfect for support, communication, events, and creativity.\n\n",
            color=0x6a95a2
        )
        embed.set_thumbnail(url="https://media.discordapp.net/attachments/1059135171540029552/1130471478526214194/1688931915684.png?width=451&height=660")
        embed.add_field(name="OWNER",value="***```                      NAO                        ```***",inline=False)
        embed.add_field(name="*Discord Server*",value=f'[<:Discord:1130420728404127844>](https://discord.gg/9qEKYTPYBH) [Discord](https://discord.gg/9qEKYTPYBH)')
        embed.add_field(name="*Instagram*",value=f'[<:instagram:1111801696545427536>](https://www.instagram.com/nao0.vt/) [Instagram](https://www.instagram.com/nao0.vt/)')
        embed.add_field(name="*Youtube*",value=f'[<:Youtube:751271269936136192>](https://www.youtube.com/@Nao0vt) [Youtube](https://www.youtube.com/@Nao0vt)')
        embed.add_field(name="*Twitch*",value=f'[<:Twitch:751271226558775377>](https://www.twitch.tv/naovtuberen) [Twitch](https://www.twitch.tv/naovtuberen)')
        embed.add_field(name="*Tiktok*",value=f'[<:tiktok:1111801698860679279>](https://www.tiktok.com/@nao0.vt) [Tiktok](https://www.tiktok.com/@nao0.vt)')
        embed.add_field(name="*Twitter*",value=f'[<:twitter:1111801706133606502>](https://twitter.com/Nao0_vt) [Twitter](https://twitter.com/Nao0_vt)')
        await interaction.followup.send(embed=embed)
        print(f"Bot_info used by  {interaction.user.name}")
        return
    
#------------------------------------------------------------------------------------------------------------------
    @app_commands.command(name="help", description="Get the list of available commands")
    @app_commands.guild_only()
    @app_commands.choices(command_name=[
    app_commands.Choice(name='Ask', value=1),
    app_commands.Choice(name='Reset', value=2),
    app_commands.Choice(name='Voice', value=3),
    app_commands.Choice(name='Imagine', value=4),
    app_commands.Choice(name='Channel', value=5),
    app_commands.Choice(name='Info', value=6),
    ])

    async def help(self, interaction: discord.Interaction, command_name: app_commands.Choice[int]=None):
        await interaction.response.defer(thinking=True)
    
        if command_name == None:
            embed = discord.Embed(
                color=0x6a95a2
            )
            
            
            embed.set_author(name ="OAN Plugins Commands" , icon_url="https://media.discordapp.net/attachments/1059135171540029552/1130471478526214194/1688931915684.png")
            embed.set_thumbnail(url="https://media.discordapp.net/attachments/1059135171540029552/1130471478526214194/1688931915684.png")
            embed.add_field(name="**Ask**",value=f'</help:1130896856256352347> Ask')
            embed.add_field(name="**Voice**",value=f'</help:1130896856256352347> Voice')
            embed.add_field(name="**Reset**",value=f'</help:1130896856256352347> Reset')
            embed.add_field(name="**Imagine**",value=f'</help:1130896856256352347> Imagine')
            embed.add_field(name="**Channel**",value=f'</help:1130896856256352347> Channel')
            embed.add_field(name="**Info**",value=f'</help:1130896856256352347> Info')
            await interaction.followup.send(embed=embed)
            print(f"Bot **Help** used by {interaction.user.name}")
            return
    
        if command_name.value == 1: #Ask
            embed = discord.Embed(title="**Ask OAN**",
                                  description=f"""</ask:1120659235026518108> ``[prompt]``\n!ask ``[prompt]``
I can generate text, and answer your questions in an informative way.\n
**Here are some specific examples of what I can do:**
```
- I can write poems, stories, scripts, musical pieces.
- I can translate languages.
- I can answer your questions in an informative way.
- I can summarize websites and articles.
- I can write code in variety of programming languages
- I can provide images, respond to voice, and return voice. 
``` """,
            color=0x6a95a2
                                )
            if interaction.user.avatar:
                embed.set_footer(text=interaction.user.name, icon_url=interaction.user.avatar.url)
            await interaction.followup.send(embed=embed)
            print(f"help **Ask** used by {interaction.user.name}")
            return
        
        if command_name.value == 2: #Reset
            embed = discord.Embed(
            title="**Reset**",
            description=f"""</reset:1130847355537739920>
```
Starts a new chat session with OAN chatbot.
```
""",
            color=0x6a95a2)

            if interaction.user.avatar:
                embed.set_footer(text=interaction.user.name, icon_url=interaction.user.avatar.url)
            await interaction.followup.send(embed=embed)
            print(f"help **Reset** used by {interaction.user.name}")
            return
        
        if command_name.value == 3: #Voice
            embed = discord.Embed(
            title="**Voice**",
            description="""</voice:1121508357749084321> ``[attached]``
```
Initiates a voice chat session and responds to an attached voice recording with a recorded voice message.
```     
""",
            color=0x6a95a2)
            if interaction.user.avatar:
                embed.set_footer(text=interaction.user.name, icon_url=interaction.user.avatar.url)
            await interaction.followup.send(embed=embed)
            print(f"help **Voice** used by {interaction.user.name}")
            return

        if command_name.value == 4: #Imagine
            embed = discord.Embed(
            title="**Imagine**",
            description="""</imagine:1121765413584392192> ``[prompt]`` ``[model]``
```
Turn your imagination into stunning visuals with AI.
```  
""",
            color=0x6a95a2)
            if interaction.user.avatar:
                embed.set_footer(text=interaction.user.name, icon_url=interaction.user.avatar.url)
            await interaction.followup.send(embed=embed)
            print(f"help **Imagine** used by {interaction.user.name}")
            return
        if command_name.value == 5: #Channel
            embed = discord.Embed(
            title="**Set Channel**",
            description="""</channel:1130847355537739918> ``[choose channel]``
```
Create a chat channel for effortless, natural conversations with AI, allowing users to interact through text or voice recordings without the need for explicit commands.
```  
""",
            color=0x6a95a2)
            if interaction.user.avatar:
                embed.set_footer(text=interaction.user.name, icon_url=interaction.user.avatar.url)
            await interaction.followup.send(embed=embed)
            print(f"help **Channel** used by  {interaction.user.name}")
            return
        if command_name.value == 6: #Info
            embed = discord.Embed(
            title="**Info**",
            description="""</info:1130847355537739919>
```
This command provides comprehensive information about the bot. It includes details about the bot's purpose, functionality, and any notable features. Additionally, it displays information about the bot's owner, developer, and the main server where it operates.
```
""",
        
            color=0x6a95a2)
            if interaction.user.avatar:
                embed.set_footer(text=interaction.user.name, icon_url=interaction.user.avatar.url)
            await interaction.followup.send(embed=embed)
            print(f"help **Info** used by {interaction.user.name}")
            return
#-----------------------------------------------------------------------------------------
    @commands.command(name='help', help='help commands')
    @app_commands.guild_only()
    async def c_help(self, ctx):
            embed = discord.Embed(
                color=0x6a95a2
            )
            embed.set_author(name ="OAN Plugins Commands" , icon_url="https://media.discordapp.net/attachments/1059135171540029552/1130471478526214194/1688931915684.png")
            embed.set_thumbnail(url="https://media.discordapp.net/attachments/1059135171540029552/1130471478526214194/1688931915684.png")
            embed.add_field(name="**Ask**",value=f'</help:1130896856256352347> Ask')
            embed.add_field(name="**Voice**",value=f'</help:1130896856256352347> Voice')
            embed.add_field(name="**Reset**",value=f'</help:1130896856256352347> Reset')
            embed.add_field(name="**Imagine**",value=f'</help:1130896856256352347> Imagine')
            embed.add_field(name="**Channel**",value=f'</help:1130896856256352347> Channel')
            embed.add_field(name="**Info**",value=f'</help:1130896856256352347> Info')
            await ctx.reply(embed=embed)
            user_name = ctx.author.name
            print(f"Bot **Help** used by {user_name}")
            return
    

async def setup(bot: commands.Bot) -> None:
    await bot.add_cog(MyCog(bot))
